import logging
from odoo import http
from odoo.http import request, Response
from io import BytesIO
from datetime import datetime

_logger = logging.getLogger(__name__)

# Column layout (8 columns):
# A=Producto | B=Inicio | C=Entrada | D=Venta | E=Vendido | F=Final | G=Costo | H=Ventas
# D=Inicio+Entrada (formula) | F=Venta-Vendido (formula) | H=Vendido*Costo (formula)
HEADERS = ['Producto', 'Inicio', 'Entrada', 'Venta', 'Vendido', 'Final', 'Costo', 'Ventas']
COL_WIDTHS = [30, 12, 12, 12, 12, 12, 14, 18]
NUM_COLS = 8


def _try_import_excel_lib():
    """Try importing Excel libraries in order of preference.
    Returns (module, lib_name) or (None, None) if none available.
    """
    for lib_name, import_path in [
        ('openpyxl', 'openpyxl'),
        ('xlsxwriter', 'xlsxwriter'),
        ('xlwt', 'xlwt'),
    ]:
        try:
            mod = __import__(import_path)
            return mod, lib_name
        except ImportError:
            continue
    return None, None


class ExcelReportController(http.Controller):

    @http.route('/pos/sales_report_excel', type='http', auth='user')
    def download_sales_excel(self, session_id=False, supplier_ids='', date_start='', date_stop='', **kwargs):
        """Generate and download Excel sales report for a POS session
        :param session_id: POS session ID (optional if date_start/date_stop provided)
        :param supplier_ids: comma-separated partner IDs to filter by supplier
        :param date_start: start date YYYY-MM-DD
        :param date_stop: end date YYYY-MM-DD
        """
        # Parse session
        session = False
        if session_id:
            session = request.env['pos.session'].browse(int(session_id))
            if not session.exists():
                return request.not_found()

        # Parse supplier_ids
        supplier_id_list = False
        if supplier_ids:
            try:
                supplier_id_list = [int(sid) for sid in supplier_ids.split(',') if sid.strip()]
                if not supplier_id_list:
                    supplier_id_list = False
            except (ValueError, AttributeError):
                supplier_id_list = False

        # Parse dates
        ds = False
        de = False
        if date_start:
            try:
                ds = datetime.strptime(date_start, '%Y-%m-%d').date()
            except ValueError:
                pass
        if date_stop:
            try:
                de = datetime.strptime(date_stop, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Need either session or dates
        if not session and not ds:
            return Response("Se requiere sesion o rango de fechas.", status=400)

        data = request.env['pos.session'].get_sales_excel_data(
            session=session,
            date_start=ds,
            date_stop=de,
            supplier_ids=supplier_id_list,
        )

        excel_lib, lib_name = _try_import_excel_lib()
        if not excel_lib:
            _logger.error("No Excel library available (openpyxl, xlsxwriter, xlwt)")
            return Response(
                "Error: No se puede generar el reporte Excel. "
                "Instale una libreria Excel (openpyxl o xlsxwriter) en el entorno Python de Odoo.",
                status=500,
            )

        try:
            if lib_name == 'openpyxl':
                buffer = self._generate_openpyxl(data)
            elif lib_name == 'xlsxwriter':
                buffer = self._generate_xlsxwriter(data)
            elif lib_name == 'xlwt':
                buffer = self._generate_xlwt(data)
        except Exception:
            _logger.exception("Error generating Excel report")
            return Response(
                "Error al generar el reporte Excel. Revise los logs del servidor.",
                status=500,
            )

        # Generate filename
        if session:
            name_part = (session.name or 'session').replace(' ', '_').replace('/', '-')
        elif ds and de:
            name_part = '%s_al_%s' % (ds.strftime('%d-%m-%Y'), de.strftime('%d-%m-%Y'))
        elif ds:
            name_part = 'desde_%s' % ds.strftime('%d-%m-%Y')
        elif de:
            name_part = 'hasta_%s' % de.strftime('%d-%m-%Y')
        else:
            name_part = datetime.now().strftime('%d-%m-%Y')

        filename = "Reporte_Ventas_%s_%s.xlsx" % (
            name_part,
            datetime.now().strftime('%d-%m-%Y'),
        )

        content_type = (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            if lib_name != 'xlwt'
            else 'application/vnd.ms-excel'
        )

        return Response(
            buffer.getvalue(),
            headers=[
                ('Content-Type', content_type),
                ('Content-Disposition', 'attachment; filename="%s"' % filename),
                ('Content-Length', str(len(buffer.getvalue()))),
            ],
            direct_passthrough=True,
        )

    # ============================================================
    # XlsxWriter generator (preferred — ships with Odoo 17)
    # ============================================================
    def _generate_xlsxwriter(self, data):
        import xlsxwriter

        buffer = BytesIO()
        wb = xlsxwriter.Workbook(buffer, {'in_memory': True})
        ws = wb.add_worksheet("Reporte de Ventas")

        # --- Styles ---
        title_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 16, 'bold': True, 'font_color': '#4A4A4A',
            'align': 'center', 'valign': 'vcenter',
        })
        subtitle_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 12, 'font_color': '#666666',
            'align': 'center', 'valign': 'vcenter',
        })
        date_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 14, 'bold': True, 'font_color': '#4A4A4A',
            'align': 'right', 'valign': 'vcenter',
        })
        header_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'bold': True, 'font_color': '#FFFFFF',
            'bg_color': '#875A7B', 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        data_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 10,
            'border': 1, 'border_color': '#CCCCCC',
        })
        data_center = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        data_right = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'align': 'right', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC', 'num_format': '#,##0.00',
        })
        final_red = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'bold': True, 'font_color': '#DC3545',
            'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        final_green = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'bold': True, 'font_color': '#28A745',
            'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        total_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'bold': True, 'font_color': '#4A4A4A',
            'bg_color': '#F2F2F2',
            'border': 1, 'border_color': '#CCCCCC',
        })
        total_center = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'bold': True, 'font_color': '#4A4A4A',
            'bg_color': '#F2F2F2', 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        total_right = wb.add_format({
            'font_name': 'Arial', 'font_size': 10, 'bold': True, 'font_color': '#4A4A4A',
            'bg_color': '#F2F2F2', 'align': 'right', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC', 'num_format': '#,##0.00',
        })
        summary_header_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 11, 'bold': True, 'font_color': '#FFFFFF',
            'bg_color': '#28A745', 'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        summary_label_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 11, 'bold': True,
            'border': 1, 'border_color': '#CCCCCC',
        })
        summary_value_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 11, 'bold': True,
            'align': 'right', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC', 'num_format': '#,##0.00',
        })
        summary_sign_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 11, 'bold': True,
            'align': 'center', 'valign': 'vcenter',
            'border': 1, 'border_color': '#CCCCCC',
        })
        footer_fmt = wb.add_format({
            'font_name': 'Arial', 'font_size': 9, 'italic': True, 'font_color': '#875A7B',
        })

        # Column widths
        col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for i, w in enumerate(COL_WIDTHS):
            ws.set_column('%s:%s' % (col_letters[i], col_letters[i]), w)

        row = 0  # xlsxwriter is 0-indexed

        # === HEADER ===
        ws.merge_range(row, 0, row, NUM_COLS - 1, "Reporte de Ventas", title_fmt)
        row += 1
        ws.merge_range(row, 0, row, NUM_COLS - 1, data.get('company_name', ''), subtitle_fmt)
        row += 1
        # Mostrar filtro de proveedor si existe
        supplier_names = data.get('supplier_filter_names', [])
        if supplier_names:
            supplier_text = "Proveedor: " + ", ".join(supplier_names)
            ws.merge_range(row, 0, row, NUM_COLS - 1, supplier_text, subtitle_fmt)
            row += 1
        ws.merge_range(row, 5, row, NUM_COLS - 1, data.get('report_date', ''), date_fmt)
        row += 2

        # === TABLE HEADERS ===
        for col, header in enumerate(HEADERS):
            ws.write(row, col, header, header_fmt)
        row += 1
        data_start_row = row

        # === DATA ROWS ===
        for product in data.get('products', []):
            # A: Producto
            ws.write_string(row, 0, product.get('name', ''), data_fmt)
            # B: Inicio
            ws.write_number(row, 1, product.get('stock_initial', 0), data_center)
            # C: Entrada
            ws.write_number(row, 2, product.get('stock_entry', 0), data_center)
            # D: Venta = Inicio + Entrada (formula)
            ws.write_formula(row, 3, '=B%d+C%d' % (row + 1, row + 1), data_center)
            # E: Vendido
            ws.write_number(row, 4, product.get('qty_sold', 0), data_center)
            # F: Final = Venta - Vendido (formula)
            ws.write_formula(row, 5, '=D%d-E%d' % (row + 1, row + 1), data_center)
            # G: Costo
            ws.write_number(row, 6, product.get('cost', 0), data_right)
            # H: Ventas = Vendido * Costo (formula)
            ws.write_formula(row, 7, '=E%d*G%d' % (row + 1, row + 1), data_right)
            row += 1

        data_end_row = row - 1

        # === TOTAL ROW with SUMA formulas ===
        total_row = row
        ws.write(row, 0, 'TOTAL', total_fmt)
        ws.write_formula(row, 1, '=SUM(B%d:B%d)' % (data_start_row + 1, data_end_row + 1), total_center)
        ws.write_formula(row, 2, '=SUM(C%d:C%d)' % (data_start_row + 1, data_end_row + 1), total_center)
        ws.write_formula(row, 3, '=SUM(D%d:D%d)' % (data_start_row + 1, data_end_row + 1), total_center)
        ws.write_formula(row, 4, '=SUM(E%d:E%d)' % (data_start_row + 1, data_end_row + 1), total_center)
        ws.write_formula(row, 5, '=SUM(F%d:F%d)' % (data_start_row + 1, data_end_row + 1), total_center)
        ws.write_blank(row, 6, total_right)  # Costo unitario no se suma
        ws.write_formula(row, 7, '=SUM(H%d:H%d)' % (data_start_row + 1, data_end_row + 1), total_right)
        row += 2

        # === SUMMARY TABLE ===
        ws.merge_range(row, 0, row, 2, "Resumen", summary_header_fmt)
        row += 1

        ws.write(row, 0, "A Entregar", summary_label_fmt)
        # A Entregar = celda H del TOTAL (columna Ventas)
        ws.write_formula(row, 1, '=H%d' % (total_row + 1), summary_value_fmt)
        ws.write_blank(row, 2, summary_value_fmt)
        row += 1

        ws.write(row, 0, "Firma:", summary_label_fmt)
        ws.merge_range(row, 1, row, 2, "_________________________", summary_sign_fmt)
        row += 2

        # === FOOTER ===
        ws.merge_range(row, 0, row, NUM_COLS - 1, "Reporte diseñado por PosJVL", footer_fmt)
        row += 1
        ws.merge_range(row, 0, row, NUM_COLS - 1, "Contacto: +5352046805", footer_fmt)

        wb.close()
        buffer.seek(0)
        return buffer

    # ============================================================
    # openpyxl generator (if available)
    # ============================================================
    def _generate_openpyxl(self, data):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        # === STYLES ===
        title_font = Font(name='Arial', size=16, bold=True, color='4A4A4A')
        subtitle_font = Font(name='Arial', size=12, color='666666')
        date_font = Font(name='Arial', size=14, bold=True, color='4A4A4A')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='875A7B', end_color='875A7B', fill_type='solid')
        data_font = Font(name='Arial', size=10)
        total_font = Font(name='Arial', size=10, bold=True, color='4A4A4A')
        total_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        summary_header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
        summary_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        summary_font = Font(name='Arial', size=11, bold=True)
        footer_font = Font(name='Arial', size=9, italic=True, color='875A7B')
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        # Column widths
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row = 1

        # === HEADER ===
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        cell = ws.cell(row=row, column=1, value="Reporte de Ventas")
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        cell = ws.cell(row=row, column=1, value=data.get('company_name', ''))
        cell.font = subtitle_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Mostrar filtro de proveedor si existe
        supplier_names = data.get('supplier_filter_names', [])
        if supplier_names:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
            cell = ws.cell(row=row, column=1, value="Proveedor: " + ", ".join(supplier_names))
            cell.font = subtitle_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=NUM_COLS)
        cell = ws.cell(row=row, column=6, value=data.get('report_date', ''))
        cell.font = date_font
        cell.alignment = Alignment(horizontal='right', vertical='center')
        row += 2

        # === TABLE HEADERS ===
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        data_start_row = row

        # === DATA ROWS ===
        for product in data.get('products', []):
            # A: Producto
            cell = ws.cell(row=row, column=1, value=product.get('name', ''))
            cell.font = data_font
            cell.alignment = left_align
            cell.border = thin_border

            # B: Inicio
            cell = ws.cell(row=row, column=2, value=product.get('stock_initial', 0))
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # C: Entrada
            cell = ws.cell(row=row, column=3, value=product.get('stock_entry', 0))
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # D: Venta = Inicio + Entrada (formula)
            cell = ws.cell(row=row, column=4, value='=B%d+C%d' % (row, row))
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # E: Vendido
            cell = ws.cell(row=row, column=5, value=product.get('qty_sold', 0))
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # F: Final = Venta - Vendido (formula)
            cell = ws.cell(row=row, column=6, value='=D%d-E%d' % (row, row))
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

            # G: Costo
            cell = ws.cell(row=row, column=7, value=product.get('cost', 0))
            cell.font = data_font
            cell.alignment = right_align
            cell.border = thin_border
            cell.number_format = '#,##0.00'

            # H: Ventas = Vendido * Costo (formula)
            cell = ws.cell(row=row, column=8, value='=E%d*G%d' % (row, row))
            cell.font = data_font
            cell.alignment = right_align
            cell.border = thin_border
            cell.number_format = '#,##0.00'

            row += 1

        data_end_row = row - 1

        # === TOTAL ROW with SUMA formulas ===
        total_row = row

        cell = ws.cell(row=row, column=1, value='TOTAL')
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = left_align

        # Inicio
        cell = ws.cell(row=row, column=2, value='=SUM(B%d:B%d)' % (data_start_row, data_end_row))
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border; cell.alignment = center_align

        # Entrada
        cell = ws.cell(row=row, column=3, value='=SUM(C%d:C%d)' % (data_start_row, data_end_row))
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border; cell.alignment = center_align

        # Venta
        cell = ws.cell(row=row, column=4, value='=SUM(D%d:D%d)' % (data_start_row, data_end_row))
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border; cell.alignment = center_align

        # Vendido
        cell = ws.cell(row=row, column=5, value='=SUM(E%d:E%d)' % (data_start_row, data_end_row))
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border; cell.alignment = center_align

        # Final
        cell = ws.cell(row=row, column=6, value='=SUM(F%d:F%d)' % (data_start_row, data_end_row))
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border; cell.alignment = center_align

        # Costo (unitario — no se suma)
        cell = ws.cell(row=row, column=7)
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border; cell.alignment = right_align

        # Ventas
        cell = ws.cell(row=row, column=8, value='=SUM(H%d:H%d)' % (data_start_row, data_end_row))
        cell.font = total_font; cell.fill = total_fill; cell.border = thin_border
        cell.alignment = right_align; cell.number_format = '#,##0.00'

        row += 2

        # === SUMMARY TABLE ===
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value="Resumen")
        cell.font = summary_header_font
        cell.fill = summary_header_fill
        cell.alignment = center_align
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = summary_header_fill
        row += 1

        cell = ws.cell(row=row, column=1, value="A Entregar")
        cell.font = summary_font
        cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        # A Entregar = celda H del TOTAL
        cell = ws.cell(row=row, column=2, value='=H%d' % total_row)
        cell.font = summary_font; cell.alignment = right_align; cell.border = thin_border
        cell.number_format = '#,##0.00'
        for c in range(2, 4):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        cell = ws.cell(row=row, column=1, value="Firma:")
        cell.font = summary_font
        cell.border = thin_border
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=2, value="_________________________")
        cell.font = summary_font; cell.alignment = center_align; cell.border = thin_border
        for c in range(2, 4):
            ws.cell(row=row, column=c).border = thin_border
        row += 2

        # === FOOTER ===
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        cell = ws.cell(row=row, column=1, value="Reporte diseñado por PosJVL")
        cell.font = footer_font
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        cell = ws.cell(row=row, column=1, value="Contacto: +5352046805")
        cell.font = footer_font

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ============================================================
    # xlwt generator (last resort — .xls format)
    # ============================================================
    def _generate_xlwt(self, data):
        import xlwt

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("Reporte de Ventas")

        # Styles
        title_style = xlwt.easyxf('font: name Arial, height 320, bold on, color 0x4A4A4A; align: horiz center, vert center')
        subtitle_style = xlwt.easyxf('font: name Arial, height 240, color 0x666666; align: horiz center, vert center')
        date_style = xlwt.easyxf('font: name Arial, height 280, bold on, color 0x4A4A4A; align: horiz right, vert center')
        header_style = xlwt.easyxf(
            'font: name Arial, height 200, bold on, color white; '
            'align: horiz center, vert center; '
            'pattern: pattern solid, fore_color 0x875A7B; '
            'borders: left thin, right thin, top thin, bottom thin'
        )
        data_style = xlwt.easyxf('font: name Arial, height 200; borders: left thin, right thin, top thin, bottom thin')
        data_center = xlwt.easyxf('font: name Arial, height 200; align: horiz center; borders: left thin, right thin, top thin, bottom thin')
        data_right = xlwt.easyxf('font: name Arial, height 200; align: horiz right; borders: left thin, right thin, top thin, bottom thin', num_format_str='#,##0.00')
        total_style = xlwt.easyxf(
            'font: name Arial, height 200, bold on, color 0x4A4A4A; '
            'align: horiz center; '
            'pattern: pattern solid, fore_color 0xF2F2F2; '
            'borders: left thin, right thin, top thin, bottom thin'
        )
        total_right_style = xlwt.easyxf(
            'font: name Arial, height 200, bold on, color 0x4A4A4A; '
            'align: horiz right; '
            'pattern: pattern solid, fore_color 0xF2F2F2; '
            'borders: left thin, right thin, top thin, bottom thin',
            num_format_str='#,##0.00'
        )
        summary_header_style = xlwt.easyxf(
            'font: name Arial, height 220, bold on, color white; '
            'align: horiz center; '
            'pattern: pattern solid, fore_color 0x28A745; '
            'borders: left thin, right thin, top thin, bottom thin'
        )
        summary_label_style = xlwt.easyxf('font: name Arial, height 220, bold on; borders: left thin, right thin, top thin, bottom thin')
        summary_value_style = xlwt.easyxf('font: name Arial, height 220, bold on; align: horiz right; borders: left thin, right thin, top thin, bottom thin', num_format_str='#,##0.00')
        summary_sign_style = xlwt.easyxf('font: name Arial, height 220, bold on; align: horiz center; borders: left thin, right thin, top thin, bottom thin')
        footer_style = xlwt.easyxf('font: name Arial, height 180, italic on, color 0x875A7B')

        # Column widths
        for i, w in enumerate(COL_WIDTHS):
            ws.col(i).width = w * 256

        row = 0

        # === HEADER ===
        ws.write_merge(row, 0, NUM_COLS - 1, "Reporte de Ventas", title_style)
        row += 1
        ws.write_merge(row, 0, NUM_COLS - 1, data.get('company_name', ''), subtitle_style)
        row += 1
        # Mostrar filtro de proveedor si existe
        supplier_names = data.get('supplier_filter_names', [])
        if supplier_names:
            ws.write_merge(row, 0, NUM_COLS - 1, "Proveedor: " + ", ".join(supplier_names), subtitle_style)
            row += 1
        ws.write_merge(row, 5, NUM_COLS - 1, data.get('report_date', ''), date_style)
        row += 2

        # === TABLE HEADERS ===
        for col, header in enumerate(HEADERS):
            ws.write(row, col, header, header_style)
        row += 1

        data_start_row = row

        # === DATA ROWS ===
        for product in data.get('products', []):
            # A: Producto
            ws.write(row, 0, product.get('name', ''), data_style)
            # B: Inicio
            ws.write(row, 1, product.get('stock_initial', 0), data_center)
            # C: Entrada
            ws.write(row, 2, product.get('stock_entry', 0), data_center)
            # D: Venta = Inicio + Entrada (formula)
            ws.write(row, 3, xlwt.Formula('B%d+C%d' % (row + 1, row + 1)), data_center)
            # E: Vendido
            ws.write(row, 4, product.get('qty_sold', 0), data_center)
            # F: Final = Venta - Vendido (formula)
            ws.write(row, 5, xlwt.Formula('D%d-E%d' % (row + 1, row + 1)), data_center)
            # G: Costo
            ws.write(row, 6, product.get('cost', 0), data_right)
            # H: Ventas = Vendido * Costo (formula)
            ws.write(row, 7, xlwt.Formula('E%d*G%d' % (row + 1, row + 1)), data_right)
            row += 1

        data_end_row = row - 1

        # === TOTAL ROW with SUMA formulas ===
        total_row = row
        ws.write(row, 0, 'TOTAL', total_style)
        ws.write(row, 1, xlwt.Formula('SUM(B%d:B%d)' % (data_start_row + 1, data_end_row + 1)), total_style)
        ws.write(row, 2, xlwt.Formula('SUM(C%d:C%d)' % (data_start_row + 1, data_end_row + 1)), total_style)
        ws.write(row, 3, xlwt.Formula('SUM(D%d:D%d)' % (data_start_row + 1, data_end_row + 1)), total_style)
        ws.write(row, 4, xlwt.Formula('SUM(E%d:E%d)' % (data_start_row + 1, data_end_row + 1)), total_style)
        ws.write(row, 5, xlwt.Formula('SUM(F%d:F%d)' % (data_start_row + 1, data_end_row + 1)), total_style)
        ws.write(row, 6, '', total_right_style)  # Costo unitario no se suma
        ws.write(row, 7, xlwt.Formula('SUM(H%d:H%d)' % (data_start_row + 1, data_end_row + 1)), total_right_style)
        row += 2

        # === SUMMARY TABLE ===
        ws.write_merge(row, 0, 2, "Resumen", summary_header_style)
        row += 1

        ws.write(row, 0, "A Entregar", summary_label_style)
        # A Entregar = celda H del TOTAL
        ws.write_merge(row, 1, 2, xlwt.Formula('H%d' % (total_row + 1)), summary_value_style)
        row += 1

        ws.write(row, 0, "Firma:", summary_label_style)
        ws.write_merge(row, 1, 2, "_________________________", summary_sign_style)
        row += 2

        # === FOOTER ===
        ws.write_merge(row, 0, NUM_COLS - 1, "Reporte diseñado por PosJVL", footer_style)
        row += 1
        ws.write_merge(row, 0, NUM_COLS - 1, "Contacto: +5352046805", footer_style)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
